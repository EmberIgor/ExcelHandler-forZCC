import openpyxl
import os
import re
from enum import Enum

from openpyxl.cell import MergedCell


class TargetFields(Enum):
    """待修改Excel表头名称"""
    idField = '申請番号'
    automaticExtractionField = '申請書に記載された内容（マクロ自動抽出）'
    dataField = 'SAP仕入先一覧'
    differencesField = '相違箇所自動判定'
    reasonField = '再鑑者コメント'
    supplierField = '仕入先'


def get_excel_list():
    """
    获取当前目录下excel文件列表
    :return: 当前目录下excel文件列表
    """
    all_item_list = os.listdir('.')
    excel_list = []
    for item in all_item_list:
        item_matcher = re.match('^(?!~\\$).+.xlsx$|^(?!~\\$).+.xlsm$', item)
        if item_matcher is not None:
            excel_list.append({
                'name': item_matcher.group(),
                'path': f'{os.getcwd()}\\{item_matcher.group()}',
            })
    return excel_list


def load_excel(excel_name):
    """
    分拣并读取excel
    :param excel_name: excel名称
    :return: excel详情
    """
    excel_detail = {
        "workBook": None,
        "type": "other",
        "excelName": excel_name,
    }
    wb = openpyxl.load_workbook(filename=excel_name)
    for sheet_item in wb.sheetnames:
        if sheet_item == '再鑑結果':
            excel_detail = load_target_excel(excel_name, wb)
            break
        elif sheet_item == '国内':
            excel_detail = load_manage_excel(excel_name, wb)
            break
    return excel_detail


def load_manage_excel(excel_name, wb):
    """
    加载管理excel数据
    :param wb: workBook
    :param excel_name: excel文件全名
    :return: excel概况、管理sheet内容
    """
    excel_detail = {
        "workBook": wb,
        "type": "manage",
        "excelName": excel_name,
        "domesticData": [],
        "foreignData": [],
        "surroundingData": [],
    }
    # 加载国内sheet
    domestic_sheet = wb.get_sheet_by_name('国内')
    # 获取国内表头
    domestic_header_list = []
    domestic_header_row = domestic_sheet[2]
    for domestic_header in domestic_header_row:
        domestic_header_list.append(domestic_header.value)
    # 获取国内数据
    for row in domestic_sheet.iter_rows(min_row=3, max_row=domestic_sheet.max_row, min_col=1,
                                        max_col=len(domestic_header_list)):
        invalid_row = True
        row_data = {}
        for cell in row:
            if cell.value is not None:
                invalid_row = False
            row_data[domestic_header_list[cell.col_idx - 1]] = cell.value
        if invalid_row:
            break
        else:
            excel_detail['domesticData'].append(row_data)
    # 加载国外sheet
    foreign_sheet = wb.get_sheet_by_name('海外')
    # 获取国外表头
    foreign_header_list = []
    foreign_header_row = foreign_sheet[2]
    for foreign_header in foreign_header_row:
        foreign_header_list.append(foreign_header.value)
    # 获取国外数据
    for row in foreign_sheet.iter_rows(min_row=3, max_row=foreign_sheet.max_row, min_col=1,
                                       max_col=len(foreign_header_list)):
        invalid_row = True
        row_data = {}
        for cell in row:
            if cell.value is not None:
                invalid_row = False
            row_data[foreign_header_list[cell.col_idx - 1]] = cell.value
        if invalid_row:
            break
        else:
            excel_detail['foreignData'].append(row_data)
    # 加载周边sheet
    surrounding_sheet = wb.get_sheet_by_name('周辺')
    # 获取周边表头
    surrounding_header_list = []
    surrounding_header_row = surrounding_sheet[1]
    for surrounding_header in surrounding_header_row:
        surrounding_header_list.append(surrounding_header.value)
    # 获取周边数据
    for row in surrounding_sheet.iter_rows(min_row=3, max_row=surrounding_sheet.max_row, min_col=1,
                                           max_col=len(surrounding_header_list)):
        invalid_row = True
        row_data = {}
        for cell in row:
            if cell.value is not None:
                invalid_row = False
            row_data[surrounding_header_list[cell.col_idx - 1]] = cell.value
        if invalid_row:
            break
        else:
            excel_detail['surroundingData'].append(row_data)
    return excel_detail


def load_target_excel(excel_name, wb):
    """
    加载待修改excel数据
    :param wb: workBook
    :param excel_name: excel文件全名
    :return: excel概况、再鑑結果sheet内容
    """
    excel_detail = {
        "workBook": wb,
        "type": "reappraisalResult",
        "excelName": excel_name,
        "reappraisalResult": load_reappraisal_result_sheet(wb),
        "extractionResult": load_extraction_request_sheet(wb),
        "requestItemResult": load_request_item_sheet(wb),

    }
    return excel_detail


def load_reappraisal_result_sheet(wb):
    """
    加载再鑑結果sheet数据
    :param wb: workbook
    :return: sheet概况、各条目信息
    """
    sheet_name = '再鑑結果'
    sheet = wb[sheet_name]
    max_row = sheet.max_row
    differences_list = []
    fields_info = {
        "automaticExtractionField": None,
        "idField": None,
        "dataField": None,
        "differencesField": None,
        "reasonField": None,
        "supplierField": None
    }
    # 下面这个循环用于获取表头坐标
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        if None not in fields_info.values():
            break
        for cell in row:
            if cell is not None and not isinstance(cell, MergedCell):
                cell_info = {
                    'column_letter': cell.column_letter,
                    'col_idx': cell.col_idx,
                    'row': cell.row
                }
                for key, value in fields_info.items():
                    if cell.value == TargetFields[key].value:
                        fields_info[key] = cell_info
                if None not in fields_info.values():
                    break
    if None in fields_info.values():
        return differences_list
    # 获取条目属性
    header_row = sheet[f"{fields_info['automaticExtractionField']['row'] + 1}"]
    raw_data_props = []
    for cell in header_row[
                fields_info['automaticExtractionField']['col_idx'] - 1:fields_info['dataField']['col_idx'] - 1]:
        raw_data_props.append(cell.value)
    data_props = []
    for cell in header_row[fields_info['dataField']['col_idx'] - 1:fields_info['differencesField']['col_idx'] - 1]:
        data_props.append(cell.value)
    differences_props = []
    for cell in header_row[fields_info['differencesField']['col_idx'] - 1:fields_info['reasonField']['col_idx'] - 1]:
        differences_props.append(cell.value)
    # 获取差异数据
    current_idx = 0
    current_row = sheet[f"{fields_info['automaticExtractionField']['row'] + 2 + current_idx}"]
    while int(fields_info['automaticExtractionField']['row'] + 2 + current_idx) <= max_row:
        if current_row[fields_info['idField']['col_idx'] - 1].value is None:
            current_idx += 1
            current_row = sheet[f"{fields_info['automaticExtractionField']['row'] + 2 + current_idx}"]
            continue
        if not re.match(r'^[\da-zA-Z]*$', str(current_row[fields_info['idField']['col_idx'] - 1].value)):
            break
        differences_item = {
            TargetFields.idField.value: str(current_row[fields_info['idField']['col_idx'] - 1].value),
            'rawData': {},
            'data': {},
            'different': [],
            'reason': '',
            'reasonCell': None
        }
        # 写入原数据
        prop_idx = 0
        for prop in raw_data_props:
            differences_item['rawData'][prop] = current_row[
                prop_idx + fields_info['automaticExtractionField']['col_idx'] - 1].value
            prop_idx += 1
        # 写入现数据
        prop_idx = 0
        for prop in data_props:
            differences_item['data'][prop] = current_row[prop_idx + fields_info['dataField']['col_idx'] - 1].value
            prop_idx += 1
        # 写入差异点
        prop_idx = 0
        for i in range(len(differences_props)):
            if current_row[prop_idx + fields_info['differencesField']['col_idx'] - 1].value == '×':
                differences_item['different'].append(
                    header_row[prop_idx + fields_info['differencesField']['col_idx'] - 1].value)
            prop_idx += 1
        # 写入reasonCell
        differences_item['reasonCell'] = current_row[fields_info['reasonField']['col_idx'] - 1]
        # 迭代
        current_idx += 1
        current_row = sheet[f"{fields_info['automaticExtractionField']['row'] + 2 + current_idx}"]
        differences_list.append(differences_item)
    return differences_list


def load_request_item_sheet(wb):
    sheet_name = 'S_LAR_87012089'
    sheet = wb[sheet_name]
    request_item_list = []
    max_row = sheet.max_row
    for i in range(3, max_row + 1):
        row = sheet[f"{i}"]
        if row[0].value is None:
            continue
        request_item = {
            "id": str(row[0].value),
            "type": str(row[1].value),
            "reason": row[2]
        }
        request_item_list.append(request_item)
    return request_item_list


def load_extraction_request_sheet(wb):
    """
    加载"申請書からの抽出結果"表
    :param wb: excel workbook
    :return: 抽出結果
    """
    sheet_name = "申請書からの抽出結果"
    sheet = wb[sheet_name]
    extraction_request_list = []
    max_row = sheet.max_row
    for i in range(2, max_row + 1):
        cell = sheet[f"A{i}"]
        if cell.value is not None:
            extraction_request_list.append(cell.value)
    return extraction_request_list
